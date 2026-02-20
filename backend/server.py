from fastapi import FastAPI, APIRouter, HTTPException
from fastapi.responses import StreamingResponse
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
from bson import ObjectId
import os
import logging
from pathlib import Path
from pydantic import BaseModel, Field
from typing import List, Optional, Dict
import uuid
from datetime import datetime
import io
import base64

# For PDF generation
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image as RLImage, PageBreak
from reportlab.lib.units import inch

# For Excel generation
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.drawing.image import Image as XLImage


ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

# MongoDB connection
mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

# Create the main app without a prefix
app = FastAPI()

# Create a router with the /api prefix
api_router = APIRouter(prefix="/api")


# Define Models
class PropertyConfig(BaseModel):
    property_name: str
    property_address: str
    property_type: str  # "apartment" or "house"
    bedrooms: int
    bathrooms: int
    balconies: int = 0
    parking_spots: int = 0
    has_kitchen: bool = True
    additional_notes: Optional[str] = None

class Property(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    config: PropertyConfig
    created_at: datetime = Field(default_factory=datetime.utcnow)

class ChecklistItemStatus(BaseModel):
    item_id: str
    category: str
    item_name: str
    status: str = "pending"  # pending, pass, fail, needs_attention
    notes: Optional[str] = None
    photos: List[str] = []  # base64 encoded images
    checked_at: Optional[datetime] = None

class Inspection(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    property_id: str
    property_config: PropertyConfig
    inspection_date: datetime = Field(default_factory=datetime.utcnow)
    status: str = "in_progress"  # in_progress, completed
    checklist_items: List[ChecklistItemStatus] = []
    completed_at: Optional[datetime] = None
    inspector_name: Optional[str] = None

class InspectionCreate(BaseModel):
    property_config: PropertyConfig
    inspector_name: Optional[str] = None

class InspectionUpdate(BaseModel):
    status: Optional[str] = None
    inspector_name: Optional[str] = None

class ChecklistItemUpdate(BaseModel):
    status: str
    notes: Optional[str] = None
    photos: Optional[List[str]] = None


# Helper function to generate checklist items based on property configuration
def generate_checklist_items(config: PropertyConfig) -> List[ChecklistItemStatus]:
    items = []
    item_counter = 0
    
    def add_item(category: str, item_name: str):
        nonlocal item_counter
        items.append(ChecklistItemStatus(
            item_id=f"item_{item_counter}",
            category=category,
            item_name=item_name,
            status="pending"
        ))
        item_counter += 1
    
    # Living Room
    add_item("Living Room", "Walls - Paint quality and finish")
    add_item("Living Room", "Walls - Check for cracks or dampness")
    add_item("Living Room", "Ceiling - Paint and finish")
    add_item("Living Room", "Ceiling - Check for leaks or stains")
    add_item("Living Room", "Flooring - Type and condition")
    add_item("Living Room", "Flooring - Check for levelness")
    add_item("Living Room", "Windows - Opening and closing mechanism")
    add_item("Living Room", "Windows - Glass condition and locks")
    add_item("Living Room", "Doors - Opening and closing smoothly")
    add_item("Living Room", "Doors - Locks and handles")
    add_item("Living Room", "Electrical - Light switches working")
    add_item("Living Room", "Electrical - Power outlets functioning")
    add_item("Living Room", "Electrical - Light fixtures installed")
    
    # Bedrooms
    for i in range(config.bedrooms):
        bedroom_name = f"Bedroom {i+1}"
        add_item(bedroom_name, "Walls - Paint quality and finish")
        add_item(bedroom_name, "Walls - Check for cracks or dampness")
        add_item(bedroom_name, "Ceiling - Paint and finish")
        add_item(bedroom_name, "Flooring - Type and condition")
        add_item(bedroom_name, "Windows - Opening/closing and locks")
        add_item(bedroom_name, "Doors - Opening/closing and locks")
        add_item(bedroom_name, "Electrical - Switches and outlets")
        add_item(bedroom_name, "Wardrobe/Closet - Doors and shelves")
    
    # Bathrooms
    for i in range(config.bathrooms):
        bathroom_name = f"Bathroom {i+1}"
        add_item(bathroom_name, "Walls - Tiles condition and grouting")
        add_item(bathroom_name, "Flooring - Tiles and drainage slope")
        add_item(bathroom_name, "Ceiling - Waterproofing and finish")
        add_item(bathroom_name, "Door - Waterproof and lock")
        add_item(bathroom_name, "Toilet - Flush mechanism working")
        add_item(bathroom_name, "Toilet - Check for leaks")
        add_item(bathroom_name, "Sink - Taps and drainage")
        add_item(bathroom_name, "Sink - Check for leaks")
        add_item(bathroom_name, "Shower/Bathtub - Water pressure")
        add_item(bathroom_name, "Shower/Bathtub - Hot and cold water")
        add_item(bathroom_name, "Shower/Bathtub - Drainage")
        add_item(bathroom_name, "Exhaust Fan - Working condition")
        add_item(bathroom_name, "Electrical - Switches and outlets")
        add_item(bathroom_name, "Mirror and Accessories - Installation")
    
    # Kitchen
    if config.has_kitchen:
        add_item("Kitchen", "Walls - Tiles condition and grouting")
        add_item("Kitchen", "Flooring - Tiles and condition")
        add_item("Kitchen", "Ceiling - Paint and finish")
        add_item("Kitchen", "Cabinets - Doors and shelves")
        add_item("Kitchen", "Cabinets - Drawer mechanisms")
        add_item("Kitchen", "Countertop - Material and finish")
        add_item("Kitchen", "Sink - Taps and drainage")
        add_item("Kitchen", "Sink - Check for leaks")
        add_item("Kitchen", "Gas Connection - Safety and working")
        add_item("Kitchen", "Chimney/Exhaust - Installation and working")
        add_item("Kitchen", "Electrical - Switches and outlets")
        add_item("Kitchen", "Windows - Ventilation")
    
    # Balconies
    for i in range(config.balconies):
        balcony_name = f"Balcony {i+1}"
        add_item(balcony_name, "Flooring - Tiles and waterproofing")
        add_item(balcony_name, "Railing - Height and stability")
        add_item(balcony_name, "Drainage - Water outlet")
        add_item(balcony_name, "Door - Opening and locking")
        add_item(balcony_name, "Electrical - Outlet availability")
    
    # Parking
    for i in range(config.parking_spots):
        parking_name = f"Parking Spot {i+1}"
        add_item(parking_name, "Space - Size and markings")
        add_item(parking_name, "Access - Entry and exit clearance")
        add_item(parking_name, "Lighting - Adequate illumination")
    
    # General/Common Areas
    add_item("Entry/Corridor", "Main Door - Lock and security features")
    add_item("Entry/Corridor", "Main Door - Finish and condition")
    add_item("Entry/Corridor", "Corridor - Paint and finish")
    add_item("Entry/Corridor", "Corridor - Lighting")
    
    # Electrical - General
    add_item("Electrical System", "Main distribution board - Installation")
    add_item("Electrical System", "MCB/Circuit breakers - Working")
    add_item("Electrical System", "Earthing - Proper grounding")
    add_item("Electrical System", "Wiring - Quality and concealment")
    
    # Plumbing - General
    add_item("Plumbing System", "Water supply - Pressure check")
    add_item("Plumbing System", "Water supply - Quality check")
    add_item("Plumbing System", "Drainage - All drains flowing properly")
    add_item("Plumbing System", "Water meter - Installation and reading")
    
    # Overall
    add_item("Overall", "Ventilation - Adequate fresh air")
    add_item("Overall", "Paint - Overall quality and finish")
    add_item("Overall", "Cleanliness - Construction debris removed")
    add_item("Overall", "Documentation - Handover documents")
    
    return items


# API Routes
@api_router.get("/")
async def root():
    return {"message": "CheckMate API - Property Inspection System"}


@api_router.post("/inspections", response_model=Inspection)
async def create_inspection(inspection_create: InspectionCreate):
    """Create a new inspection with generated checklist"""
    # Generate checklist items based on property configuration
    checklist_items = generate_checklist_items(inspection_create.property_config)
    
    inspection = Inspection(
        property_id=str(uuid.uuid4()),
        property_config=inspection_create.property_config,
        checklist_items=checklist_items,
        inspector_name=inspection_create.inspector_name
    )
    
    await db.inspections.insert_one(inspection.model_dump())
    return inspection


@api_router.get("/inspections", response_model=List[Inspection])
async def get_all_inspections():
    """Get all inspections"""
    inspections = await db.inspections.find().to_list(1000)
    return [Inspection(**inspection) for inspection in inspections]


@api_router.get("/inspections/{inspection_id}", response_model=Inspection)
async def get_inspection(inspection_id: str):
    """Get a specific inspection by ID"""
    inspection = await db.inspections.find_one({"id": inspection_id})
    if not inspection:
        raise HTTPException(status_code=404, detail="Inspection not found")
    return Inspection(**inspection)


@api_router.put("/inspections/{inspection_id}", response_model=Inspection)
async def update_inspection(inspection_id: str, update: InspectionUpdate):
    """Update inspection metadata"""
    inspection = await db.inspections.find_one({"id": inspection_id})
    if not inspection:
        raise HTTPException(status_code=404, detail="Inspection not found")
    
    update_data = update.model_dump(exclude_unset=True)
    
    if update.status == "completed":
        update_data["completed_at"] = datetime.utcnow()
    
    await db.inspections.update_one(
        {"id": inspection_id},
        {"$set": update_data}
    )
    
    updated_inspection = await db.inspections.find_one({"id": inspection_id})
    return Inspection(**updated_inspection)


@api_router.put("/inspections/{inspection_id}/items/{item_id}", response_model=Inspection)
async def update_checklist_item(inspection_id: str, item_id: str, update: ChecklistItemUpdate):
    """Update a specific checklist item"""
    inspection = await db.inspections.find_one({"id": inspection_id})
    if not inspection:
        raise HTTPException(status_code=404, detail="Inspection not found")
    
    # Find and update the specific item
    checklist_items = inspection.get("checklist_items", [])
    item_found = False
    
    for item in checklist_items:
        if item["item_id"] == item_id:
            item["status"] = update.status
            if update.notes is not None:
                item["notes"] = update.notes
            if update.photos is not None:
                item["photos"] = update.photos
            item["checked_at"] = datetime.utcnow()
            item_found = True
            break
    
    if not item_found:
        raise HTTPException(status_code=404, detail="Checklist item not found")
    
    await db.inspections.update_one(
        {"id": inspection_id},
        {"$set": {"checklist_items": checklist_items}}
    )
    
    updated_inspection = await db.inspections.find_one({"id": inspection_id})
    return Inspection(**updated_inspection)


@api_router.delete("/inspections/{inspection_id}")
async def delete_inspection(inspection_id: str):
    """Delete an inspection"""
    result = await db.inspections.delete_one({"id": inspection_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Inspection not found")
    return {"message": "Inspection deleted successfully"}


@api_router.get("/inspections/{inspection_id}/pdf")
async def generate_pdf_report(inspection_id: str):
    """Generate and download PDF report for an inspection"""
    inspection = await db.inspections.find_one({"id": inspection_id})
    if not inspection:
        raise HTTPException(status_code=404, detail="Inspection not found")
    
    inspection_obj = Inspection(**inspection)
    
    # Create PDF in memory
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    elements = []
    styles = getSampleStyleSheet()
    
    # Title
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=24,
        textColor=colors.HexColor('#2563eb'),
        spaceAfter=30,
        alignment=1  # Center
    )
    elements.append(Paragraph("CheckMate - Property Inspection Report", title_style))
    elements.append(Spacer(1, 0.2*inch))
    
    # Property Details
    prop_config = inspection_obj.property_config
    property_data = [
        ["Property Name:", prop_config.property_name],
        ["Address:", prop_config.property_address],
        ["Property Type:", prop_config.property_type.title()],
        ["Bedrooms:", str(prop_config.bedrooms)],
        ["Bathrooms:", str(prop_config.bathrooms)],
        ["Balconies:", str(prop_config.balconies)],
        ["Parking Spots:", str(prop_config.parking_spots)],
        ["Inspection Date:", inspection_obj.inspection_date.strftime("%Y-%m-%d %H:%M")],
        ["Inspector:", inspection_obj.inspector_name or "N/A"],
        ["Status:", inspection_obj.status.upper()],
    ]
    
    prop_table = Table(property_data, colWidths=[2*inch, 4*inch])
    prop_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#e5e7eb')),
        ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ('TOPPADDING', (0, 0), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -1), 1, colors.grey),
    ]))
    elements.append(prop_table)
    elements.append(Spacer(1, 0.3*inch))
    
    # Summary Statistics
    total_items = len(inspection_obj.checklist_items)
    passed = sum(1 for item in inspection_obj.checklist_items if item.status == "pass")
    failed = sum(1 for item in inspection_obj.checklist_items if item.status == "fail")
    needs_attention = sum(1 for item in inspection_obj.checklist_items if item.status == "needs_attention")
    pending = sum(1 for item in inspection_obj.checklist_items if item.status == "pending")
    
    elements.append(Paragraph("Summary Statistics", styles['Heading2']))
    summary_data = [
        ["Total Items", "Passed", "Failed", "Needs Attention", "Pending"],
        [str(total_items), str(passed), str(failed), str(needs_attention), str(pending)]
    ]
    summary_table = Table(summary_data, colWidths=[1.2*inch]*5)
    summary_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2563eb')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 10),
        ('TOPPADDING', (0, 0), (-1, -1), 10),
        ('GRID', (0, 0), (-1, -1), 1, colors.grey),
    ]))
    elements.append(summary_table)
    elements.append(Spacer(1, 0.3*inch))
    
    # Checklist Items by Category
    elements.append(Paragraph("Detailed Checklist", styles['Heading2']))
    elements.append(Spacer(1, 0.1*inch))
    
    # Group items by category
    categories = {}
    for item in inspection_obj.checklist_items:
        if item.category not in categories:
            categories[item.category] = []
        categories[item.category].append(item)
    
    for category, items in categories.items():
        elements.append(Paragraph(f"<b>{category}</b>", styles['Heading3']))
        
        category_data = [["Item", "Status", "Notes"]]
        for item in items:
            status_color = {
                "pass": "GREEN",
                "fail": "RED",
                "needs_attention": "ORANGE",
                "pending": "GRAY"
            }.get(item.status, "GRAY")
            
            notes_text = item.notes[:100] + "..." if item.notes and len(item.notes) > 100 else (item.notes or "-")
            
            category_data.append([
                item.item_name,
                f"<font color='{status_color}'><b>{item.status.upper()}</b></font>",
                notes_text
            ])
        
        cat_table = Table(category_data, colWidths=[2.5*inch, 1*inch, 2.5*inch])
        cat_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#f3f4f6')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ('TOPPADDING', (0, 0), (-1, -1), 6),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ]))
        elements.append(cat_table)
        elements.append(Spacer(1, 0.2*inch))
    
    # Build PDF
    doc.build(elements)
    buffer.seek(0)
    
    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={
            "Content-Disposition": f"attachment; filename=inspection_{inspection_id}.pdf"
        }
    )


@api_router.get("/inspections/{inspection_id}/excel")
async def generate_excel_report(inspection_id: str):
    """Generate and download Excel report for an inspection"""
    inspection = await db.inspections.find_one({"id": inspection_id})
    if not inspection:
        raise HTTPException(status_code=404, detail="Inspection not found")
    
    inspection_obj = Inspection(**inspection)
    
    # Create Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Inspection Report"
    
    # Header styling
    header_fill = PatternFill(start_color="2563eb", end_color="2563eb", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=12)
    
    # Title
    ws.merge_cells('A1:F1')
    ws['A1'] = "CheckMate - Property Inspection Report"
    ws['A1'].font = Font(bold=True, size=16, color="2563eb")
    ws['A1'].alignment = Alignment(horizontal='center')
    
    # Property Details
    row = 3
    prop_config = inspection_obj.property_config
    ws[f'A{row}'] = "Property Name:"
    ws[f'A{row}'].font = Font(bold=True)
    ws[f'B{row}'] = prop_config.property_name
    row += 1
    
    ws[f'A{row}'] = "Address:"
    ws[f'A{row}'].font = Font(bold=True)
    ws[f'B{row}'] = prop_config.property_address
    row += 1
    
    ws[f'A{row}'] = "Property Type:"
    ws[f'A{row}'].font = Font(bold=True)
    ws[f'B{row}'] = prop_config.property_type.title()
    row += 1
    
    ws[f'A{row}'] = "Configuration:"
    ws[f'A{row}'].font = Font(bold=True)
    ws[f'B{row}'] = f"{prop_config.bedrooms} BR, {prop_config.bathrooms} BA, {prop_config.balconies} Balconies, {prop_config.parking_spots} Parking"
    row += 1
    
    ws[f'A{row}'] = "Inspection Date:"
    ws[f'A{row}'].font = Font(bold=True)
    ws[f'B{row}'] = inspection_obj.inspection_date.strftime("%Y-%m-%d %H:%M")
    row += 1
    
    ws[f'A{row}'] = "Inspector:"
    ws[f'A{row}'].font = Font(bold=True)
    ws[f'B{row}'] = inspection_obj.inspector_name or "N/A"
    row += 2
    
    # Summary Statistics
    total_items = len(inspection_obj.checklist_items)
    passed = sum(1 for item in inspection_obj.checklist_items if item.status == "pass")
    failed = sum(1 for item in inspection_obj.checklist_items if item.status == "fail")
    needs_attention = sum(1 for item in inspection_obj.checklist_items if item.status == "needs_attention")
    pending = sum(1 for item in inspection_obj.checklist_items if item.status == "pending")
    
    ws[f'A{row}'] = "Summary Statistics"
    ws[f'A{row}'].font = Font(bold=True, size=12)
    row += 1
    
    summary_headers = ["Total Items", "Passed", "Failed", "Needs Attention", "Pending"]
    for col, header in enumerate(summary_headers, start=1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
    row += 1
    
    summary_values = [total_items, passed, failed, needs_attention, pending]
    for col, value in enumerate(summary_values, start=1):
        cell = ws.cell(row=row, column=col)
        cell.value = value
        cell.alignment = Alignment(horizontal='center')
    row += 2
    
    # Checklist Items
    ws[f'A{row}'] = "Detailed Checklist"
    ws[f'A{row}'].font = Font(bold=True, size=12)
    row += 1
    
    # Headers
    headers = ["Category", "Item", "Status", "Notes", "Photos", "Checked At"]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
    row += 1
    
    # Data rows
    for item in inspection_obj.checklist_items:
        ws.cell(row=row, column=1, value=item.category)
        ws.cell(row=row, column=2, value=item.item_name)
        
        status_cell = ws.cell(row=row, column=3, value=item.status.upper())
        if item.status == "pass":
            status_cell.font = Font(color="008000", bold=True)
        elif item.status == "fail":
            status_cell.font = Font(color="FF0000", bold=True)
        elif item.status == "needs_attention":
            status_cell.font = Font(color="FFA500", bold=True)
        
        ws.cell(row=row, column=4, value=item.notes or "-")
        ws.cell(row=row, column=5, value=f"{len(item.photos)} photo(s)" if item.photos else "No photos")
        ws.cell(row=row, column=6, value=item.checked_at.strftime("%Y-%m-%d %H:%M") if item.checked_at else "-")
        row += 1
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 40
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 18
    
    # Save to buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename=inspection_{inspection_id}.xlsx"
        }
    )


# Include the router in the main app
app.include_router(api_router)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()
